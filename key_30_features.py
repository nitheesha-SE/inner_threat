import xlrd
import xlsxwriter
import openpyxl
import os, sys
import math
import collections
from xlutils.copy import copy
def heading(worksheet):
    worksheet.write(0, 0, "time sum left down")
    worksheet.write(0, 1, "maxtime left  down")
    worksheet.write(0, 2, "mintime left down")
    worksheet.write(0, 3, "std_dev left click down")
    worksheet.write(0, 4, "count left down count")
    worksheet.write(0, 5, "time sum left release ")
    worksheet.write(0, 6, "maxtime left  release")
    worksheet.write(0, 7, "mintime left  release")
    worksheet.write(0, 8, "std_dev left release")
    worksheet.write(0, 9, "count left release")
    worksheet.write(0, 10, " left time diff sum")
    worksheet.write(0, 11, "max left click diff")
    worksheet.write(0, 12, "min left click diff")
    worksheet.write(0, 13, "std_dev left_click diff")
    worksheet.write(0, 14, "left down and release count")
    worksheet.write(0, 15, "time sum  Right  down ")
    worksheet.write(0, 16, "maxtime right  down")
    worksheet.write(0, 17, "mintime right  down")
    worksheet.write(0, 18, "std_dev right  down")
    worksheet.write(0, 19, "count right down")
    worksheet.write(0, 20, " Right click release")
    worksheet.write(0, 21, "count right release")
    worksheet.write(0, 22, " right time diff")
    worksheet.write(0, 23, "right difference count")
    worksheet.write(0, 24, " sum Keypress time")
    worksheet.write(0, 25, "count keypress time")
    worksheet.write(0, 26, "time sum numpad  ")
    worksheet.write(0, 27, "count numpad press")
    worksheet.write(0, 28, " keyspeed  time")
    worksheet.write(0, 29, " backspace count")
    worksheet.write(0, 30, "value")
    
def leftclickdown(worksheet,sheet,nrows):

    # while(k < nrows):
    # print "rownum",nrows
    # col_list = []
    sum = 0.0000
    average = 0.0000
    count = 0
    col_list = []
    sqr = 0.0000
    # col_data = sheet.cell(rownum,7).value
	#loop for each row 
    for rownum in range(1, nrows):
		#check if the movement is left click down
        if ((sheet.cell(rownum, 4).value).strip() == "Left Click Down"):
			#store the time value from the sheet
            time1 = sheet.cell(rownum, 3).value
			#add the values to the list
            col_list.append(time1)
            sum = sum + time1
            count = count + 1
            # print sum
    if count == 0:
        count = 1
	#find the average
    average = sum / count
    if (len(col_list) == 0):
        min = 0
        max = 0
    else:
        min = col_list[0]
        max = col_list[0]
	#find max and min values int the list
    for col_values in col_list:
        if (col_values >= max):
            max = col_values
        if (col_values <= min):
            min = col_values
        sqr += ((col_values - average) ** 2)
    std_dev = math.sqrt(sqr / count)
    worksheet.write(1, 0, sum)
    worksheet.write(1, 1, max)
    worksheet.write(1, 2, min)
    worksheet.write(1, 3, std_dev)
    worksheet.write(1, 4, count)
def leftclickrelease(worksheet,sheet,nrows):

    sqr = 0
    sum = 0.0000
    average = 0.0000
    count = 0
    col_list = []
    for rownum in range(1, nrows):
        # col_data = sheet.cell(rownum,7).value
        if ((sheet.cell(rownum, 4).value).strip() == "Left Click Release"):
            time1 = sheet.cell(rownum, 3).value
            sum = sum + time1
            col_list.append(time1)
            count = count + 1
    # print sum
    if count == 0:
        count = 1
    average = sum / count
    if (len(col_list) == 0):
        max = 0
        min = 0
    else:
        max = col_list[0]
        min = col_list[0]
    for col_values in col_list:
        if col_values > max:
            max = col_values
        if col_values < min:
            min = col_values
        sqr += ((col_values - average) ** 2)
    std_dev = math.sqrt(sqr / count)
    worksheet.write(1, 5, sum)
    worksheet.write(1, 6, max)
    worksheet.write(1, 7, min)
    worksheet.write(1, 8, std_dev)
    worksheet.write(1, 9, count)
def mouseleftclick(worksheet,sheet,nrows):
    j = 1
    k = 200
    p = 1
    sum = 0.0000
    average = 0.0000
    count = 0
    sqr = 0
    col_list = []
    for rownum in range(1, nrows - 1):
        # col_data = sheet.cell(rownum,7).value
        if ((sheet.cell(rownum, 4).value).strip() == "Left Click Down" and (
                sheet.cell(rownum + 1, 4).value).strip() == "Left Click Release"):
            time1 = sheet.cell(rownum, 3).value
            time2 = sheet.cell(rownum + 1, 3).value
            timediff = abs(time1 - time2)
            col_list.append(timediff)
            sum = sum + timediff
            count = count + 1
    if count == 0:
        count = 1
    # print sum
    average = sum / count
    if (len(col_list) == 0):
        max = 0
        min = 0
    else:
        max = col_list[0]
        min = col_list[0]
    for col_values in col_list:
        if col_values > max:
            max = col_values
        if col_values < min:
            min = col_values
        sqr += ((col_values - average) ** 2)
    std_dev = math.sqrt(sqr / count)
    worksheet.write(1, 10, sum)
    worksheet.write(1, 11, max)
    worksheet.write(1, 12, min)
    worksheet.write(1, 13, std_dev)
    worksheet.write(1, 14, count)
def rightclickdown(worksheet,sheet,nrows):

    sum = 0.0000
    average = 0.0000
    count = 0
    col_list = []
    for rownum in range(1, nrows):
        # col_data = sheet.cell(rownum,7).value
        if ((sheet.cell(rownum, 4).value).strip() == "Right Click Down"):
            time1 = sheet.cell(rownum, 3).value
            col_list.append(time1)
            sum = sum + time1
            count = count + 1
            # print sum
    if count == 0:
        count = 1
    average = sum / count
    if (len(col_list) == 0):
        max = 0
        min = 0
    else:
        max = col_list[0]
        min = col_list[0]
    for col_values in col_list:
        if col_values > max:
            max = col_values
        if col_values < min:
            min = col_values
        sqr += ((col_values - average) ** 2)
    std_dev = math.sqrt(sqr / count)
    worksheet.write(1, 15, sum)
    worksheet.write(1, 16, max)
    worksheet.write(1, 17, min)
    worksheet.write(1, 18, std_dev)
    worksheet.write(1, 19, count)
def rightclickrelease(worksheet,sheet,nrows):

    # print "rownum",nrows
    # col_list = []
    sum = 0.0000
    average = 0.0000
    count = 0
    for rownum in range(1, nrows):
        # col_data = sheet.cell(rownum,7).value
        if ((sheet.cell(rownum, 4).value).strip() == "Right Click Release"):
            time1 = sheet.cell(rownum, 3).value
            sum = sum + time1
            count = count + 1
            # print sum
    if count == 0:
        count = 1
    average = sum / count
    worksheet.write(1, 20, sum)
    worksheet.write(1, 21, count)
def mouserightclick(worksheet,sheet,nrows):

    sum = 0.0000
    average = 0.0000
    count = 0
    for rownum in range(1, nrows - 1):
        # col_data = sheet.cell(rownum,7).value
        if ((sheet.cell(rownum, 4).value).strip() == "Right Click Down" and (
        sheet.cell(rownum + 1, 4).value).strip() == "Right Click Release"):
            time1 = sheet.cell(rownum, 3).value
            time2 = sheet.cell(rownum + 1, 3).value
            timediff = abs(time1 - time2)
            sum = sum + timediff
            count = count + 1
    # print sum
    if count == 0:
        count = 1
    average = sum / count
    worksheet.write(1, 22, sum)
    worksheet.write(1, 23, count)
def keypresstime(worksheet,sheet,nrows):

    sum = 0.0000
    average = 0.0000
    count = 0
    for rownum in range(1, nrows):
        # col_data = sheet.cell(rownum,7).value
        if ((sheet.cell(rownum, 4).value).strip()[:8] == "Keypress" and (sheet.cell(rownum, 4).value).strip()[
                                                                        :15] != "Keypress numpad"):
            time1 = sheet.cell(rownum, 3).value
            sum = sum + time1
            count = count + 1
    # print sum
    if count == 0:
        count = 1
    average = sum / count
    worksheet.write(1, 24, sum)
    worksheet.write(1, 25, count)
def numpadkey(worksheet,sheet,nrows):

    sum = 0.0000
    average = 0.0000
    count = 0
    for rownum in range(1, nrows):
        # col_data = sheet.cell(rownum,7).value
        if ((sheet.cell(rownum, 4).value).strip()[:15] == "Keypress numpad"):
            time1 = sheet.cell(rownum, 3).value
            sum = sum + time1
            count = count + 1
            # prnt sum
    if count == 0:
        count = 1
    average = sum / count
    worksheet.write(1, 26, sum)
    worksheet.write(1, 27, count)
def keyspeed(worksheet,sheet,nrows):

    sum = 0.0000
    average = 0.0000
    count = 0
    j = 1
    for rownum in range(j, nrows):
        if ((sheet.cell(rownum, 4).value).strip()[:8] == "Keypress"):
            while ((sheet.cell(rownum + 1, 4).value).strip()[:8] == "Keypress"):
                time1 = sheet.cell(rownum, 3).value
                time2 = sheet.cell(rownum + 1, 3).value
                timediff = abs(time1 - time2)
                sum = sum + timediff
                rownum = rownum + 1
                count = count + 1
    if count == 0:
        count = 1
    # average = sum/count
    worksheet.write(1, 28, sum)
    j = j + 1
def backspacecount(worksheet,sheet,nrows,result):

    count = 0
    average = 0.0000
    j = 1
    for rownum in range(j, nrows):
        if ((sheet.cell(rownum, 4).value).strip() == "Keypress backspace"):
            count = count + 1
    worksheet.write(1, 29, count)
    if (result[1] == '1' or result[1] == '3' or result[1] == '2' or result[1] == '4'):
        worksheet.write(1, 30, 1)
    elif (result[1] == '5' or result[1] == '6'):
        worksheet.write(1, 30, 2)
def processFile(filename, output):
    # result = output
    result = os.path.splitext(output)[0]
    # print result
    wb = xlrd.open_workbook(filename)
    sheet = wb.sheet_by_index(0)
    # print "opened file:",filename
    final = "C:\\new_key_feature\\new_key_feature_R5\\"
    path = final + result + ".xlsx"
    workbook2 = xlsxwriter.Workbook(path)
    worksheet = workbook2.add_worksheet("Sheet 1")
    nrows = sheet.nrows
    heading(worksheet)
    # capture mouse left click down  time 
    leftclickdown(worksheet,sheet,nrows)
    # capture mouse left click down  time 
    leftclickrelease(worksheet,sheet,nrows)
    # capture mouse click time 
    mouseleftclick(worksheet,sheet,nrows)
    # capture mouse left click down  time 
    rightclickdown(worksheet,sheet,nrows)
    # capture mouse right click down  time 
    rightclickrelease(worksheet,sheet,nrows)
    # capture mouse right click time 
    mouserightclick(worksheet,sheet,nrows)
    # count of keypress and sum of the time	
    keypresstime(worksheet,sheet,nrows)
    # count of numberpad keys
    numpadkey(worksheet,sheet,nrows)
    # capture key value time
    keyspeed(worksheet,sheet,nrows)
    # capture backspace
    backspacecount(worksheet,sheet,nrows,result)
    workbook2.close()


address = "C:\\R5\\"
directory = os.listdir(address)
for filename in directory:
    output = filename
    filename = address + filename
    print(filename)
    processFile(filename, output)



    # address = "C:\\R1\\R1009.xlsx"
    # output = "R1009-s.xlsx"
    # # # # # # #filename = address+filename
    # print(address)
    # processFile(address ,output)