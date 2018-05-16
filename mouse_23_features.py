import xlrd
import xlsxwriter
import openpyxl
import os, sys
import math
import collections
from xlutils.copy import copy
def heading(worksheet):
    worksheet.write(0, 0, " speed sum")
    worksheet.write(0, 1, " speed max")
    worksheet.write(0, 2, " speed min")
    worksheet.write(0, 3, "speed_std_dev")
    worksheet.write(0, 4, "total speed")
    worksheet.write(0, 5, " dist sum")
    worksheet.write(0, 6, " dist max")
    worksheet.write(0, 7, " dist min")
    worksheet.write(0, 8, "dist_std_dev")
    worksheet.write(0, 9, "direct_distance")
    worksheet.write(0, 10, "distance_diff")
    worksheet.write(0, 11, " x-axis average")
    worksheet.write(0, 12, " x-axis max")
    worksheet.write(0, 13, " x-axis min")
    worksheet.write(0, 14, "x_axis_std_dev")
    worksheet.write(0, 15, " y-axis average")
    worksheet.write(0, 16, " y-axis max")
    worksheet.write(0, 17, " y-axis min")
    worksheet.write(0, 18, "y_axis_std_dev")
    worksheet.write(0, 19, " direction")
    worksheet.write(0, 20, " max frequency")
    worksheet.write(0, 21, " change in direction")
    worksheet.write(0, 22, "final direction")
    worksheet.write(0, 23, "value")



def speedfind(worksheet,sheet,nrows,result):
    l = 2
    k = nrows
    p = 1
	#loop for all the rows
    while (l <= nrows):
        col_list = []
        speed_sum = 0.0000
        average = 0.0000
        dis_sum = 0.0000
        time_sum = 0.0000
        j = l
        distance_list = []
        time_list = []
        # for loop in the range of next num after 800 to last row
        for rownum in range(j, k):
            # any time greater than 800 breaks the for loop
            # add the speed time to col_list array
            col_data = sheet.cell(rownum, 7).value
            col_list.append(col_data)
            # add all the distance values in the given group to distance_list
            dis_data = sheet.cell(rownum, 6).value
            distance_list.append(dis_data)
            # add all the time values in the given group to time_list
            time_data = sheet.cell(rownum, 3).value
            time_list.append(time_data)
            # increment each row if the time is not  greater than 700
            l = l + 1
			#check if the cell value is greater than 800 or if it is not mouse moement
            if ((sheet.cell(rownum, 3).value) > 800 or (sheet.cell(rownum, 4).value).strip() != "Mouse Movement"):
                if (sheet.cell(rownum, 4).value).strip() != "Mouse Movement":
                    col_list.remove(col_data)
                l = l - 1
                break
        # find the sum of all the speed values in the list
        # print time_list
        for col_values in col_list:
            if col_values == '#DIV/0!':
                col_values = 0
            if col_values == None:
                col_values = 0
            speed_sum = speed_sum + col_values
		#check if the col_list list is empty and give default values
        if len(col_list) == 0:
            lenght = 1
            max = 0
            min = 0
        # find the average of the values by dividing sum to the totla num of values in the list
        else:
            lenght = len(col_list)
            max = col_list[0]
            min = col_list[0]
        average = speed_sum / lenght
        sqr = 0.0000
        # find the max and min of all the values present in the list
        for col_values in col_list:
            if col_values == '#DIV/0!':
                col_values = 0
            if col_values == None:
                col_values = 0
            if (col_values >= max):
                max = col_values
            if (col_values <= min):
                min = col_values
            # find the standard deveation
            sqr += sqr + ((col_values - average) ** 2)
        if math.isinf(float(sqr)):
            # print 'isnan'
            sqr = 0
        std_dev = math.sqrt(sqr / lenght)
        # print the values in the sheet
        for dis_values in distance_list:
            dis_sum = dis_sum + dis_values
        for time_values in time_list:
            time_sum = time_sum + time_values
        if time_sum == 0:
            time_sum = 1
        speed_avg = dis_sum / time_sum
		#find the sheet name and write the value accordingly
        if (len(col_list) == 0):
            if (result[1] == '1' or result[1] == '3' or result[1] == '2' or result[1] == '4'):
                worksheet.write(p, 23, 0)
            elif (result[1] == '5' or result[1] == '6'):
                worksheet.write(p, 23, 0)
        else:
            if (result[1] == '1' or result[1] == '3' or result[1] == '2' or result[1] == '4'):
                worksheet.write(p, 23, 1)
            elif (result[1] == '5' or result[1] == '6'):
                worksheet.write(p, 23, 2)
        # print ',sqr',sqr
        # print 'lenght',lenght
        # print std_dev
        worksheet.write(p, 0, speed_sum)
        worksheet.write(p, 1, max)
        worksheet.write(p, 2, min)
        worksheet.write(p, 3, std_dev)
        worksheet.write(p, 4, speed_avg)
        p = p + 1
        l = l + 1
def distancefind(worksheet,sheet,nrows):
    l = 2
    k = nrows
    p = 1

    while (l < nrows - 1):
        distance_list = []
        sum = 0.0000
        average = 0.0000
        j = l
        x1 = sheet.cell(l - 1, 1).value
        y1 = sheet.cell(l - 1, 2).value
        # print 'x1',x1
        # print'y1',y1
        # for loop in the range of next num after 700 to last row
        for rownum in range(j, k - 1):
            dis_data = sheet.cell(rownum, 6).value
            distance_list.append(dis_data)
            l = l + 1
            # any time greater than 700 breaks the for loop
            if ((sheet.cell(rownum, 3).value) > 800 or (sheet.cell(rownum, 4).value).strip() != "Mouse Movement"):
                l = l - 1
                break
                # add all the distance values in the given group to distance_list
                # increment each row if the time is not greater than 700
        # print 'last value',sheet.cell(l-1,3).value
        x2 = sheet.cell(l, 1).value
        y2 = sheet.cell(l, 2).value
        # print 'x2',x2
        # print 'y2',y2

        # print distance_list
        for col_values in distance_list:
            if col_values == '#DIV/0!':
                col_values = 0
            if col_values == None:
                col_values = 0
            sum = sum + col_values
        # print 'col_values',col_values
        if len(distance_list) == 0:
            max = 0
            min = 0
        else:
            max = distance_list[0]
            min = distance_list[0]
        # print 'sum',sum
        average = sum / 100

        sqr = 0.0000
        for col_values in distance_list:
            if col_values == None:
                col_values = 0
            if (col_values >= max):
                max = col_values
            if (col_values <= min):
                min = col_values
            sqr += ((col_values - average) ** 2)
        std_dev = math.sqrt(sqr / 100)
        short_distance = math.sqrt((x1 - x2) ** 2 + (y1 - y2) ** 2)
        distance_difference = sum - short_distance
        if distance_difference < 0.08:
            distance_difference = 0.0000
        # print 'short_distance ' ,short_distance
        # print ' distance difference',distance_difference
        # print max
        # print min
        # print sqr
        worksheet.write(p, 5, sum)
        worksheet.write(p, 6, max)
        worksheet.write(p, 7, min)
        worksheet.write(p, 8, std_dev)
        worksheet.write(p, 9, short_distance)
        worksheet.write(p, 10, distance_difference)
        p = p + 1
        l = l + 1
def xaxisfind(worksheet,sheet,nrows):
    l = 2
    k = nrows
    p = 1
    while (l <= nrows):
        col_list = []
        sum = 0.0000
        average = 0.0000
        j = l
        # col_list.append(sheet.cell(l-1,9).value)
        for rownum in range(j, k):
            col_data = sheet.cell(rownum, 9).value
            col_list.append(col_data)
            l = l + 1
            if ((sheet.cell(rownum, 3).value) > 800 or (sheet.cell(rownum, 4).value).strip() != "Mouse Movement"):
                l = l - 1
                break
        # print "rownum",j
        for col_values in col_list:
            if col_values == '#DIV/0!':
                col_values = 0
            if col_values == None:
                col_values = 0
            # print col_values
            sum = sum + col_values
        # print "sum",sum
        # print sum
        if (len(col_list) == 0):
            lenght = 1
            max = 0
            min = 0
        else:
            lenght = len(col_list)
            max = col_list[0]
            min = col_list[0]
        average = sum / lenght
        sqr = 0.0000
        for col_values in col_list:
            if col_values == None:
                col_values = 0
            if col_values == None:
                col_values = 0
            if (col_values >= max):
                max = col_values
            if (col_values <= min):
                min = col_values
            sqr += ((col_values - average) ** 2)
        std_dev = math.sqrt(sqr / lenght)
        # print max
        # print min
        # print std_dev
        worksheet.write(p, 11, average)
        worksheet.write(p, 12, max)
        worksheet.write(p, 13, min)
        worksheet.write(p, 14, std_dev)

        p = p + 1
        l = l + 1

def yaxisfind(worksheet,sheet,nrows):
    l = 2
    k = nrows
    p = 1

    while (l <= nrows):
        col_list = []
        sum = 0.0000
        average = 0.0000
        j = l
        for rownum in range(j, k):
            col_data = sheet.cell(rownum, 10).value
            col_list.append(col_data)
            l = l + 1
            if ((sheet.cell(rownum, 3).value) > 800 or (sheet.cell(rownum, 4).value).strip() != "Mouse Movement"):
                l = l - 1
                break

        for col_values in col_list:
            if col_values == '#DIV/0!':
                col_values = 0
            if col_values == None:
                col_values = 0
            sum = sum + col_values
        # print sum
        if (len(col_list) == 0):
            lenght = 1
            max = 0
            min = 0
        else:
            lenght = len(col_list)
            max = col_list[0]
            min = col_list[0]
        average = sum / lenght
        sqr = 0.0000
        for col_values in col_list:
            if col_values == None:
                col_values = 0
            if (col_values >= max):
                max = col_values
            if (col_values <= min):
                min = col_values
            sqr += ((col_values - average) ** 2)
        std_dev = math.sqrt(sqr / 100)
        # print max
        # print min
        # print std_dev
        worksheet.write(p, 15, average)
        worksheet.write(p, 16, max)
        worksheet.write(p, 17, min)
        worksheet.write(p, 18, std_dev)

        p = p + 1
        l = l + 1
def directionfind(worksheet,sheet,nrows):
    l = 2
    k = nrows
    p = 1
    r = 0

    while (l <= nrows):
		#create a dictionary to save the count of directions
        frequency = collections.defaultdict(int)
        j = l
        count = 1.0000
		#loop for each row
        for rownum in range(j, k):
			#get the column value
            col_data = sheet.cell(rownum, 8).value
			#increment the column value in the dictionary
            frequency[col_data] += 1
            l = l + 1
			#check if the cell value is greater than 800 or not a mouse movement
            if ((sheet.cell(rownum, 3).value) > 800 or (sheet.cell(rownum, 4).value).strip() != "Mouse Movement"):
                l = l - 1
                break
            count = count + 1
        max = 0
        dirctn = 0
        # print frequency
        # print max
        if len(frequency) == 1 and count == 1:
            max = 0
            dirctn = 0
        # print frequency[1]
		#find the max moved direction by comparing the frequency count
        else:
            for i in range(1, 10):
                if max < frequency[i]:
                    max = frequency[i]
                    max = max / count
                    dirctn = i

        worksheet.write(p, 19, dirctn)
        worksheet.write(p, 20, max)

        l += 1
        p += 1
#final changed direction
def changedirection(worksheet,sheet,nrows):
    l = 2
    k = nrows
    p = 1
    r = 0

	#loop for eah row
    while (l <= nrows - 1):
        col_list = []
        count = 1.0000
        frequency = 0
        j = l
		#store the 1st x and y axis values
        x1 = sheet.cell(l - 1, 1).value
        y1 = sheet.cell(l - 1, 2).value
        for rownum in range(j, k - 1):
			#save the direction values to list
            col_data = sheet.cell(rownum, 8).value
            col_list.append(col_data)
            l = l + 1
			#check if cell value is greater than 800 or not a mouse movement
            if ((sheet.cell(rownum, 3).value) > 800 or (sheet.cell(rownum, 4).value).strip() != "Mouse Movement"):
                l = l - 1
                break

            count = count + 1
        # print col_list
        # print l
		#store the final x and y axis values
        x2 = sheet.cell(l, 1).value
        y2 = sheet.cell(l, 2).value
        if count == 0:
            count = 1
        for i in range(0, len(col_list) - 1):
            p1 = col_list[i]
            p2 = col_list[i + 1]
            if (p1 != p2):
                frequency = frequency + 1
        # print 'frequency',frequency
        frequency = frequency / count
        # get the direction b/w first and last pixel in the group
        if (x2 > x1 and y2 == y1):
            dir = 1
        elif (x2 > x1 and y2 > y1):
            dir = 2
        elif (x2 == x1 and y2 > y1):
            dir = 3
        elif (x2 < x1 and y2 > y1):
            dir = 4
        elif (x2 < x1 and y2 == y1):
            dir = 5
        elif (x2 < x1 and y2 < y1):
            dir = 6
        elif (x2 == x1 and y2 < y1):
            dir = 7
        elif (x2 > x1 and y2 < y1):
            dir = 8
        else:
            dir = 0

        # print 'direction',dir
        worksheet.write(p, 21, frequency)
        worksheet.write(p, 22, dir)
        l += 1
        p += 1
def processFile(filename, output):
    # result = output
    result = os.path.splitext(output)[0]
    print result
    wb = xlrd.open_workbook(filename)
    sheet = wb.sheet_by_index(0)
    # print "opened file:",filename
    final = "C:\\new_mouse_feature\\800_threhold\\new_mouse_feature_R6\\"
    path = final + result + ".xlsx"
    workbook2 = xlsxwriter.Workbook(path)
    worksheet = workbook2.add_worksheet("Sheet 1")
    nrows = sheet.nrows
	#write the first columns
    heading(worksheet)
    # speed attributes
    speedfind(worksheet,sheet,nrows,result)
    # distance
    distancefind(worksheet,sheet,nrows)
    # x-axis attributes
    xaxisfind(worksheet,sheet,nrows)
    # y-axis
    yaxisfind(worksheet,sheet,nrows)
    # direction with max times repeated
    directionfind(worksheet,sheet,nrows)
    # count of change in direction
    changedirection(worksheet,sheet,nrows)


    workbook2.close()


address = "C:\\R6\\"
directory = os.listdir(address)
for filename in directory:
    output = filename
    filename = address + filename
    print(filename)
    processFile(filename, output)

# address = "C:\\R2\\R20028.xlsx"
# output = "R1003.xlsx"
# print(address)
# processFile(address ,output)